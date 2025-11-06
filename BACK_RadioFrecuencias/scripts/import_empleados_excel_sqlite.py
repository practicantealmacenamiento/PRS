#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script mínimo para importar 'cedula' y 'nombre' desde la hoja de Excel "Base de datos"
a una tabla SQL (ejemplo con SQLite). Hace upsert manual (si existe -> UPDATE, si no -> INSERT).

Requisitos:
    pip install openpyxl
Uso:
    python scripts/import_empleados_excel_sqlite.py \
        --excel "BASE DE DATOS A&T.xlsx" \
        --db db.sqlite3 \
        --table empleados
"""
import argparse
import sqlite3
from typing import Optional, Tuple
from openpyxl import load_workbook

SINONIMOS_CEDULA = {"cedula", "cédula", "documento", "doc", "cc", "dni", "identificacion", "identificación", "id"}
SINONIMOS_NOMBRE = {"nombre", "nombres", "nombre completo", "apellidos y nombres", "empleado", "colaborador"}

def _norm_header(s: Optional[str]) -> str:
    return (str(s or "")).strip().lower()

def _digits(s: Optional[str]) -> str:
    if s is None:
        return ""
    return "".join(ch for ch in str(s).strip() if ch.isdigit())

def _strclean(s: Optional[str]) -> str:
    return str(s or "").strip()

def _find_indices(ws) -> Tuple[int, int]:
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [_norm_header(h) for h in header]

    idx_ced = -1
    idx_nom = -1
    for i, h in enumerate(headers):
        if h in SINONIMOS_CEDULA and idx_ced == -1:
            idx_ced = i
        if h in SINONIMOS_NOMBRE and idx_nom == -1:
            idx_nom = i
    if idx_ced == -1 or idx_nom == -1:
        raise ValueError("No se encontraron columnas para cédula y nombre en la primera fila.")
    return idx_ced, idx_nom

def upsert_empleado(cur, table: str, cedula: str, nombre: str):
    cur.execute(f"SELECT 1 FROM {table} WHERE cedula = ?", (cedula,))
    exists = cur.fetchone() is not None
    if exists:
        cur.execute(f"UPDATE {table} SET nombre = ? WHERE cedula = ?", (nombre, cedula))
        return "updated"
    else:
        # Agrega 'activo' como 1 (True)
        cur.execute(f"INSERT INTO {table} (cedula, nombre, activo) VALUES (?, ?, ?)", (cedula, nombre, 1))
        return "inserted"

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", required=True, help="Ruta al .xlsx (hoja: 'Base de datos')")
    ap.add_argument("--db", default="db.sqlite3", help="Ruta al SQLite db (por defecto db.sqlite3)")
    ap.add_argument("--table", default="empleados", help="Nombre de la tabla destino (por defecto 'empleado')")
    ap.add_argument("--sheet", default="Base de datos", help="Nombre de la hoja (por defecto 'Base de datos')")
    args = ap.parse_args()

    wb = load_workbook(args.excel, read_only=True, data_only=True)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"La hoja '{args.sheet}' no existe. Hojas disponibles: {', '.join(wb.sheetnames)}")
    ws = wb[args.sheet]

    idx_ced, idx_nom = _find_indices(ws)

    conn = sqlite3.connect(args.db)
    cur = conn.cursor()

    inserted = updated = skipped = 0

    try:
        # Transacción para acelerar I/O
        cur.execute("BEGIN")
        for row in ws.iter_rows(min_row=2, values_only=True):
            ced = _digits(row[idx_ced] if idx_ced < len(row) else None)
            nom = _strclean(row[idx_nom] if idx_nom < len(row) else None)
            if not ced or not nom:
                skipped += 1
                continue

            result = upsert_empleado(cur, args.table, ced, nom)
            if result == "inserted":
                inserted += 1
            elif result == "updated":
                updated += 1
            else:
                skipped += 1

        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    print("✅ Importación lista.")
    print(f"  Insertados : {inserted}")
    print(f"  Actualizados: {updated}")
    print(f"  Omitidos   : {skipped}")

if __name__ == "__main__":
    main()
